import openpyxl
import os

workbook = openpyxl.load_workbook('example.xlsx')
# print(type(workbook))
# print(workbook.get_sheet_names())
sheet = workbook.get_sheet_by_name('Sheet1')
cell = sheet['A1']
print(cell.value)
print(sheet.cell(row = 1 , column = 2))

for i in range(1, 8):
    print(i, sheet.cell(row = i, column = 2 ).value)

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'
wb.save('example2.xlsx')
sheet2 = wb.create_sheet()
sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names())
wb.save('example3.xlsx')
wb.create_sheet(index = 0, title='My other Sheet')
wb.save('example4.xlsx')